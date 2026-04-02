import openpyxl

wb = openpyxl.load_workbook('/home/node/.openclaw/media/inbound/c65676e5-96ef-4e17-8d0d-10bf7bea64e7.xlsx')
print('Sheets:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    print(f'\n=== {sheet_name} ===')
    sheet = wb[sheet_name]
    for row in sheet.iter_rows():
        row_data = [str(cell.value) if cell.value is not None else '' for cell in row]
        print('\t'.join(row_data))
